import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from data.risk_register import get_risks_with_scores, CONTROL_FACTORS

# ----------------------
# PAGE CONFIG & CSS
# ----------------------
st.set_page_config(page_title="Risk & Compliance Dashboard", page_icon="🛡️", layout="wide")

st.markdown("""
    <style>
    body, .stApp { background: #0d1117 !important; color: #fff !important; }
    .main, .block-container { background: #0d1117 !important; }
    .risk-badge-high { background: #FF0000; color: #fff; border-radius: 16px; padding: 2px 14px; font-weight: bold; }
    .risk-badge-medium { background: #FFD700; color: #222; border-radius: 16px; padding: 2px 14px; font-weight: bold; }
    .risk-badge-low { background: #70AD47; color: #fff; border-radius: 16px; padding: 2px 14px; font-weight: bold; }
    .card { background: #161b22; border-radius: 16px; box-shadow: 0 2px 12px #00000033; padding: 18px 24px; margin-bottom: 18px; }
    .metric-card { display: flex; flex-direction: column; align-items: center; justify-content: center; background: #161b22; border-radius: 16px; box-shadow: 0 2px 12px #00000033; padding: 18px 0; margin-bottom: 18px; }
    .stMetric { color: #fff !important; }
    .stDataFrame { background: #161b22 !important; }
    .stExpander { background: #161b22 !important; }
    .stSlider { color: #fff !important; }
    .stSelectbox, .stMultiSelect { color: #fff !important; }
    .stButton { color: #fff !important; }
    </style>
""", unsafe_allow_html=True)

# ----------------------
# LOAD DATA
# ----------------------
risks = get_risks_with_scores()
df = pd.DataFrame(risks)

# ----------------------
# HEADER
# ----------------------
st.markdown("<h1>🛡️ Risk & Compliance Dashboard — FY 2025</h1>", unsafe_allow_html=True)
st.markdown("<div style='font-size:1.1rem;'>Organization: <b>ACME Corp</b> | Scope: <b>IT & Cybersecurity</b> | Period: <b>FY 2025</b> | Prepared by: <b>Internal Audit</b></div>", unsafe_allow_html=True)
st.markdown(f"<div style='color:#aaa;font-size:0.95rem;'>Last updated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}</div>", unsafe_allow_html=True)

# ----------------------
# SIDEBAR FILTERS
# ----------------------
st.sidebar.title("🔧 Dashboard Controls")

risk_rating_options = ["All", "High", "Medium", "Low"]
category_options = sorted(df['category'].unique())
department_options = sorted(df['department'].unique())
status_options = ["All", "Open", "In Progress", "Closed"]

risk_rating_filter = st.sidebar.selectbox("Risk Rating Filter", risk_rating_options)
category_filter = st.sidebar.multiselect("Category Filter", category_options, default=category_options)
department_filter = st.sidebar.multiselect("Department / Owner Filter", department_options, default=department_options)
status_filter = st.sidebar.selectbox("Remediation Status", status_options)
min_risk_score = st.sidebar.slider("Minimum Risk Score", 1, 25, 1)

# Metrics
total_risks = len(df)
open_items = df[df['remediation_status'].isin(["Open", "In Progress"])].shape[0]
today = pd.Timestamp(datetime.datetime.now().date())
df['remediation_deadline_dt'] = pd.to_datetime(df['remediation_deadline'])
df['is_overdue'] = (df['remediation_deadline_dt'] < today) & (df['remediation_status'] != "Closed")
overdue_items = df['is_overdue'].sum()
st.sidebar.divider()
st.sidebar.metric("Total Risks", total_risks)
st.sidebar.metric("Open Items", open_items)
st.sidebar.metric("Overdue Items", overdue_items)

# ----------------------
# RISK SCORE CALCULATOR
# ----------------------
st.sidebar.divider()
st.sidebar.markdown("**Risk Score Calculator**")
calc_likelihood = st.sidebar.slider("Likelihood", 1, 5, 3, key="calc_likelihood")
calc_impact = st.sidebar.slider("Impact", 1, 5, 3, key="calc_impact")
calc_inherent = calc_likelihood * calc_impact
if calc_inherent >= 15:
    calc_inherent_rating = "High"
elif calc_inherent >= 8:
    calc_inherent_rating = "Medium"
else:
    calc_inherent_rating = "Low"
calc_control = st.sidebar.selectbox("Control Effectiveness", list(CONTROL_FACTORS.keys()), key="calc_control")
calc_residual = round(calc_inherent * CONTROL_FACTORS[calc_control], 1)
if calc_residual >= 15:
    calc_residual_rating = "High"
elif calc_residual >= 8:
    calc_residual_rating = "Medium"
else:
    calc_residual_rating = "Low"
st.sidebar.metric("Inherent Score", calc_inherent, calc_inherent_rating)
st.sidebar.metric("Residual Score", calc_residual, calc_residual_rating)

# ----------------------
# FILTER DATAFRAME
# ----------------------
df_filtered = df.copy()
if risk_rating_filter != "All":
    df_filtered = df_filtered[df_filtered['residual_rating'] == risk_rating_filter]
if status_filter != "All":
    df_filtered = df_filtered[df_filtered['remediation_status'] == status_filter]
df_filtered = df_filtered[df_filtered['category'].isin(category_filter)]
df_filtered = df_filtered[df_filtered['department'].isin(department_filter)]
df_filtered = df_filtered[df_filtered['residual_score'] >= min_risk_score]

# ----------------------
# TABS
# ----------------------
tabs = st.tabs([
    "🏠 Executive Dashboard",
    "📋 Risk Register",
    "🔥 Risk Heat Map",
    "📈 Risk Analytics",
    "✅ Remediation Tracker",
    "📤 Export & Reports"
])

# ----------------------
# TAB 1: EXECUTIVE DASHBOARD
# ----------------------
with tabs[0]:
    col1, col2, col3, col4 = st.columns(4)
    # Dummy prior period values for deltas
    prior_total = 12
    prior_high = 4
    prior_med = 6
    prior_low = 2
    high_count = df[df['residual_rating'] == "High"].shape[0]
    med_count = df[df['residual_rating'] == "Medium"].shape[0]
    low_count = df[df['residual_rating'] == "Low"].shape[0]
    col1.metric("Total Risks", total_risks, total_risks - prior_total)
    col2.metric("High Risk", high_count, high_count - prior_high, delta_color="inverse")
    col3.metric("Medium Risk", med_count, med_count - prior_med, delta_color="inverse")
    col4.metric("Low Risk", low_count, low_count - prior_low, delta_color="normal")

    col5, col6, col7 = st.columns(3)
    open_rem = df[df['remediation_status'].isin(["Open", "In Progress"])].shape[0]
    avg_residual = round(df['residual_score'].mean(), 2)
    pct_mitigated = round((df[df['remediation_status'] == "Closed"].shape[0] / total_risks) * 100, 1)
    col5.metric("Open Remediations", open_rem)
    col6.metric("Avg Residual Risk Score", avg_residual)
    col7.metric("% Risks Mitigated", pct_mitigated)

    # Donut chart: Risk Distribution by Residual Rating
    donut_data = df['residual_rating'].value_counts().reindex(["High", "Medium", "Low"]).fillna(0)
    fig_donut = go.Figure(data=[go.Pie(
        labels=donut_data.index,
        values=donut_data.values,
        hole=0.6,
        marker=dict(colors=["#FF0000", "#FFD700", "#70AD47"])
    )])
    fig_donut.update_layout(title="Risk Distribution by Residual Rating", showlegend=True, height=350, margin=dict(t=40, b=0))
    st.plotly_chart(fig_donut, use_container_width=True)

    # Top 5 Highest Residual Risk Items
    top5 = df.sort_values("residual_score", ascending=False).head(5)
    fig_bar = px.bar(top5, x="residual_score", y="title", orientation="h", color="residual_rating",
                     color_discrete_map={"High": "#FF0000", "Medium": "#FFD700", "Low": "#70AD47"},
                     title="Top 5 Highest Residual Risk Items", height=350)
    st.plotly_chart(fig_bar, use_container_width=True)

    # Risks by Department (stacked) - Plotly
    dept_counts = df.groupby(["department", "residual_rating"]).size().unstack(fill_value=0).reset_index()
    fig_dept = go.Figure()
    for rating, color in zip(["High", "Medium", "Low"], ["#FF0000", "#FFD700", "#70AD47"]):
        if rating in dept_counts.columns:
            fig_dept.add_trace(go.Bar(
                x=dept_counts['department'],
                y=dept_counts[rating] if rating in dept_counts else [0]*len(dept_counts),
                name=rating,
                marker_color=color
            ))
    fig_dept.update_layout(barmode='stack', title="Risks by Department (stacked by High/Medium/Low)", xaxis_title="Department", yaxis_title="Count", height=350)
    st.plotly_chart(fig_dept, use_container_width=True)

    # Risk Heat Map
    heatmap = np.zeros((5,5))
    for l in range(1,6):
        for i in range(1,6):
            score = l*i
            if score >= 15:
                heatmap[5-l,i-1] = 2  # High
            elif score >= 8:
                heatmap[5-l,i-1] = 1  # Medium
            else:
                heatmap[5-l,i-1] = 0  # Low
    fig_heat = go.Figure()
    fig_heat.add_trace(go.Heatmap(
        z=heatmap,
        x=[1,2,3,4,5],
        y=[5,4,3,2,1],
        colorscale=[[0, "#70AD47"], [0.5, "#FFD700"], [1, "#FF0000"]],
        showscale=False,
        hoverinfo="skip"
    ))
    for _, r in df.iterrows():
        fig_heat.add_trace(go.Scatter(
            x=[r['impact']], y=[6-r['likelihood']],
            mode="markers+text",
            marker=dict(size=18, color="#fff", line=dict(width=2, color="#222")),
            text=[r['risk_id']],
            textposition="middle center",
            hovertemplate=f"{r['title']}<br>Inherent: {r['inherent_score']}<br>Residual: {r['residual_score']}"
        ))
    fig_heat.update_layout(
        title="Risk Heat Map",
        xaxis_title="Impact (1–5)",
        yaxis_title="Likelihood (1–5)",
        yaxis=dict(autorange="reversed"),
        height=400,
        margin=dict(t=40, b=0)
    )
    st.plotly_chart(fig_heat, use_container_width=True)

    # Compliance Frameworks Reference
    with st.expander("Compliance Frameworks Reference"):
        frameworks = ["ISO 27001", "NIST CSF", "SOC 2", "GDPR", "PCI DSS"]
        mapping = [
            ["RSK-001", ["ISO 27001", "NIST CSF", "SOC 2"]],
            ["RSK-002", ["ISO 27001", "NIST CSF", "SOC 2", "PCI DSS"]],
            ["RSK-003", ["ISO 27001", "NIST CSF", "SOC 2"]],
            ["RSK-004", ["ISO 27001", "NIST CSF", "GDPR"]],
            ["RSK-005", ["ISO 27001", "NIST CSF", "SOC 2", "PCI DSS"]],
            ["RSK-006", ["ISO 27001", "NIST CSF"]],
            ["RSK-007", ["ISO 27001", "NIST CSF", "SOC 2"]],
            ["RSK-008", ["GDPR"]],
            ["RSK-009", ["ISO 27001", "NIST CSF", "SOC 2"]],
            ["RSK-010", ["ISO 27001", "NIST CSF", "SOC 2"]],
            ["RSK-011", ["ISO 27001", "NIST CSF"]],
            ["RSK-012", ["ISO 27001", "NIST CSF", "SOC 2", "PCI DSS"]],
        ]
        fw_table = pd.DataFrame([{
            "Risk ID": m[0], **{fw: ("✔️" if fw in m[1] else "") for fw in frameworks}
        } for m in mapping])
        st.dataframe(fw_table, use_container_width=True)

# ----------------------
# TAB 2: RISK REGISTER
# ----------------------
with tabs[1]:
    st.markdown("### 📋 Risk Register")
    def rating_badge(r):
        if r == "High":
            return f'<span class="risk-badge-high">High</span>'
        elif r == "Medium":
            return f'<span class="risk-badge-medium">Medium</span>'
        else:
            return f'<span class="risk-badge-low">Low</span>'
    def status_badge(s):
        color = {"Open": "#FF0000", "In Progress": "#FFD700", "Closed": "#70AD47"}.get(s, "#fff")
        return f'<span style="background:{color};color:#222;border-radius:12px;padding:2px 10px;font-weight:bold;">{s}</span>'
    styled_df = df_filtered.copy()
    styled_df['Inherent Rating'] = styled_df['inherent_rating'].apply(rating_badge)
    styled_df['Residual Rating'] = styled_df['residual_rating'].apply(rating_badge)
    styled_df['Status'] = styled_df['remediation_status'].apply(status_badge)
    display_cols = [
        'risk_id', 'category', 'title', 'likelihood', 'impact', 'inherent_score', 'Inherent Rating',
        'control_effectiveness', 'residual_score', 'Residual Rating', 'risk_owner', 'Status', 'remediation_deadline'
    ]
    st.markdown(styled_df[display_cols].to_html(escape=False, index=False), unsafe_allow_html=True)
    for _, r in styled_df.iterrows():
        with st.expander(f"{r['risk_id']} — {r['title']}"):
            st.write(f"**Description:** {r['description']}")
            st.write(f"**Current Controls:** {r['current_controls']}")
            st.write(f"**Remediation Action:** {r['remediation_action']}")
            st.write(f"**Dates:** Identified: {r['date_identified']} | Last Reviewed: {r['last_reviewed']}")

# ----------------------
# TAB 3: RISK HEAT MAP
# ----------------------
with tabs[2]:
    st.markdown("### 🔥 Risk Heat Map")
    fig_heat2 = go.Figure()
    fig_heat2.add_trace(go.Heatmap(
        z=heatmap,
        x=[1,2,3,4,5],
        y=[5,4,3,2,1],
        colorscale=[[0, "#70AD47"], [0.5, "#FFD700"], [1, "#FF0000"]],
        showscale=False,
        hoverinfo="skip"
    ))
    for _, r in df.iterrows():
        fig_heat2.add_trace(go.Scatter(
            x=[r['impact']], y=[6-r['likelihood']],
            mode="markers+text",
            marker=dict(size=18, color="#fff", line=dict(width=2, color="#222")),
            text=[r['risk_id']],
            textposition="middle center",
            hovertemplate=f"{r['title']}<br>Inherent: {r['inherent_score']}<br>Residual: {r['residual_score']}<br>Owner: {r['risk_owner']}<br>Category: {r['category']}"
        ))
    fig_heat2.update_layout(
        title="Risk Heat Map",
        xaxis_title="Impact (1–5)",
        yaxis_title="Likelihood (1–5)",
        yaxis=dict(autorange="reversed"),
        height=500,
        margin=dict(t=40, b=0)
    )
    st.plotly_chart(fig_heat2, use_container_width=True)
    # Scoring methodology
    st.markdown("#### Scoring Methodology")
    st.markdown("""
    | Likelihood | Description |
    |---|---|
    | 1 | Rare (<5%) |
    | 2 | Unlikely (5–25%) |
    | 3 | Possible (25–50%) |
    | 4 | Likely (50–75%) |
    | 5 | Almost Certain (>75%) |
    
    | Impact | Description |
    |---|---|
    | 1 | Negligible |
    | 2 | Minor |
    | 3 | Moderate |
    | 4 | Significant |
    | 5 | Critical |
    """)

# ----------------------
# TAB 4: RISK ANALYTICS
# ----------------------
with tabs[3]:
    st.markdown("### 📈 Risk Analytics")
    # Inherent vs Residual Risk Score per risk
    fig_scores = go.Figure()
    fig_scores.add_trace(go.Bar(
        x=df['title'], y=df['inherent_score'], name="Inherent", marker_color="#1F3864"
    ))
    fig_scores.add_trace(go.Bar(
        x=df['title'], y=df['residual_score'], name="Residual", marker_color="#FFD700"
    ))
    fig_scores.update_layout(barmode='group', title="Inherent vs Residual Risk Score", xaxis_tickangle=-30)
    st.plotly_chart(fig_scores, use_container_width=True)
    # Likelihood vs Impact scatter
    fig_scatter = px.scatter(df, x="likelihood", y="impact", size="residual_score", color="category", hover_name="title", title="Likelihood vs Impact (Bubble = Residual Score)")
    st.plotly_chart(fig_scatter, use_container_width=True)
    # Control Effectiveness distribution
    ctrl_counts = df['control_effectiveness'].value_counts().reset_index()
    ctrl_counts.columns = ['Control Effectiveness', 'Count']
    fig_ctrl = px.bar(
        ctrl_counts,
        x='Control Effectiveness',
        y='Count',
        color='Control Effectiveness',
        title="Control Effectiveness Distribution",
        labels={'Control Effectiveness':'Control Effectiveness', 'Count':'Count'}
    )
    st.plotly_chart(fig_ctrl, use_container_width=True)
    # Risks by Category (stacked)
    cat_counts = df.groupby(["category", "residual_rating"]).size().unstack(fill_value=0)
    st.bar_chart(cat_counts)
    # Remediation Status breakdown
    fig_status = px.pie(df, names='remediation_status', title="Remediation Status Breakdown", color='remediation_status', color_discrete_map={"Open": "#FF0000", "In Progress": "#FFD700", "Closed": "#70AD47"})
    st.plotly_chart(fig_status, use_container_width=True)
    # KPI: Average risk reduction %
    avg_reduction = round(np.mean((df['inherent_score'] - df['residual_score']) / df['inherent_score'] * 100), 1)
    st.metric("Average Risk Reduction %", avg_reduction)
    # Trend line chart (dummy data)
    trend_months = ["2025-01", "2025-02", "2025-03", "2025-04", "2025-05", "2025-06"]
    trend_scores = [38, 36, 34, 32, 30, 28]
    fig_trend = go.Figure()
    fig_trend.add_trace(go.Scatter(x=trend_months, y=trend_scores, mode='lines+markers', line=dict(color="#1F3864", width=3)))
    fig_trend.update_layout(title="Overall Risk Score Trend (Dummy Data)", xaxis_title="Month", yaxis_title="Total Risk Score")
    st.plotly_chart(fig_trend, use_container_width=True)

# ----------------------
# TAB 5: REMEDIATION TRACKER
# ----------------------
with tabs[4]:
    st.markdown("### ✅ Remediation Tracker")
    colA, colB, colC, colD = st.columns(4)
    open_count = df_filtered[df_filtered['remediation_status'] == "Open"].shape[0]
    inprog_count = df_filtered[df_filtered['remediation_status'] == "In Progress"].shape[0]
    closed_count = df_filtered[df_filtered['remediation_status'] == "Closed"].shape[0]
    overdue_count = df_filtered[df_filtered['is_overdue']].shape[0]
    colA.metric("Open", open_count)
    colB.metric("In Progress", inprog_count)
    colC.metric("Closed", closed_count)
    colD.metric("Overdue", overdue_count)
    # Gantt-style chart
    gantt_df = df_filtered.copy()
    gantt_df['start'] = pd.to_datetime(gantt_df['date_identified'])
    gantt_df['finish'] = pd.to_datetime(gantt_df['remediation_deadline'])
    gantt_df['color'] = gantt_df['remediation_status'].map({"Open": "#FF0000", "In Progress": "#FFD700", "Closed": "#70AD47"})
    fig_gantt = go.Figure()
    for idx, row in gantt_df.iterrows():
        fig_gantt.add_trace(go.Bar(
            x=[(row['finish'] - row['start']).days],
            y=[row['title']],
            base=[row['start']],
            orientation='h',
            marker=dict(color=row['color']),
            name=row['remediation_status'],
            hovertemplate=f"{row['risk_id']}<br>{row['title']}<br>Status: {row['remediation_status']}<br>Deadline: {row['remediation_deadline']}"
        ))
    fig_gantt.update_layout(title="Remediation Timeline", barmode='stack', xaxis_title="Date", yaxis_title="Risk Title", showlegend=False)
    st.plotly_chart(fig_gantt, use_container_width=True)
    # Table
    table_df = gantt_df[['risk_id', 'title', 'risk_owner', 'department', 'remediation_deadline', 'remediation_status']].copy()
    table_df['Days Until Deadline'] = (gantt_df['remediation_deadline_dt'] - today).dt.days
    def due_color(days, status):
        if status == "Closed":
            return "#70AD47"
        elif days < 0:
            return "#FF0000"
        elif days <= 30:
            return "#FFD700"
        else:
            return "#70AD47"
    table_df['Due Status'] = [due_color(d, s) for d, s in zip(table_df['Days Until Deadline'], table_df['remediation_status'])]
    def due_badge(days, status):
        color = due_color(days, status)
        if status == "Closed":
            return f'<span style="background:{color};color:#222;border-radius:12px;padding:2px 10px;font-weight:bold;">On Track</span>'
        elif days < 0:
            return f'<span style="background:{color};color:#fff;border-radius:12px;padding:2px 10px;font-weight:bold;">Overdue</span>'
        elif days <= 30:
            return f'<span style="background:{color};color:#222;border-radius:12px;padding:2px 10px;font-weight:bold;">Due Soon</span>'
        else:
            return f'<span style="background:{color};color:#fff;border-radius:12px;padding:2px 10px;font-weight:bold;">On Track</span>'
    table_df['Due'] = [due_badge(d, s) for d, s in zip(table_df['Days Until Deadline'], table_df['remediation_status'])]
    st.markdown(table_df[['risk_id', 'title', 'risk_owner', 'department', 'remediation_deadline', 'remediation_status', 'Days Until Deadline', 'Due']].to_html(escape=False, index=False), unsafe_allow_html=True)
    for _, r in gantt_df.iterrows():
        if r['remediation_status'] != "Closed":
            with st.expander(f"{r['risk_id']} — {r['title']}"):
                st.write(f"**Remediation Action:** {r['remediation_action']}")
                st.checkbox("Mark as Resolved", value=False, key=f"resolved_{r['risk_id']}")

# ----------------------
# TAB 6: EXPORT & REPORTS
# ----------------------
with tabs[5]:
    st.markdown("### 📤 Export & Reports")
    # Export to Excel
    def to_excel_bytes():
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Risk Register"
        headers = list(df.columns)
        ws1.append(headers)
        for row in df.itertuples(index=False):
            ws1.append(list(row))
        # Style header
        header_fill = PatternFill("solid", fgColor="1F3864")
        for cell in ws1[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
        # Color ratings
        rating_col = headers.index('residual_rating')+1
        for row in ws1.iter_rows(min_row=2, min_col=rating_col, max_col=rating_col):
            for cell in row:
                if cell.value == "High":
                    cell.fill = PatternFill("solid", fgColor="FF0000")
                elif cell.value == "Medium":
                    cell.fill = PatternFill("solid", fgColor="FFD700")
                elif cell.value == "Low":
                    cell.fill = PatternFill("solid", fgColor="70AD47")
        # Borders
        thin = Side(border_style="thin", color="888888")
        for row in ws1.iter_rows():
            for cell in row:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        # Freeze top row
        ws1.freeze_panes = "A2"
        # Auto column width
        for col in ws1.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            ws1.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2
        # Sheet 2: Heat Map Data
        ws2 = wb.create_sheet("Heat Map Data")
        ws2.append(["Likelihood\\Impact"] + [1,2,3,4,5])
        for l in range(1,6):
            row = [l]
            for i in range(1,6):
                score = l*i
                if score >= 15:
                    color = "High"
                elif score >= 8:
                    color = "Medium"
                else:
                    color = "Low"
                row.append(color)
            ws2.append(row)
        # Sheet 3: Remediation Tracker
        ws3 = wb.create_sheet("Remediation Tracker")
        ws3.append(["Risk ID", "Title", "Owner", "Department", "Deadline", "Status"])
        for _, r in df.iterrows():
            ws3.append([r['risk_id'], r['title'], r['risk_owner'], r['department'], r['remediation_deadline'], r['remediation_status']])
        # Sheet 4: Executive Summary
        ws4 = wb.create_sheet("Executive Summary")
        ws4.append(["KPI", "Value"])
        ws4.append(["Total Risks", total_risks])
        ws4.append(["High Risk", high_count])
        ws4.append(["Medium Risk", med_count])
        ws4.append(["Low Risk", low_count])
        ws4.append(["Open Remediations", open_rem])
        ws4.append(["Avg Residual Risk Score", avg_residual])
        ws4.append(["% Risks Mitigated", pct_mitigated])
        ws4.append(["Average Risk Reduction %", avg_reduction])
        ws4.append(["Last Updated", datetime.datetime.now().strftime('%Y-%m-%d %H:%M')])
        ws4.append(["Summary", "Overall risk posture is improving with ongoing remediation and control enhancements. High risks are being addressed as a priority. Medium and low risks are monitored and managed within risk appetite."])
        # Save to bytes
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        return bio
    st.download_button("📥 Download Full Risk Register (Excel)", data=to_excel_bytes(), file_name="Risk_Register_FY2025.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    st.info("PDF export: Use your browser's Print > Save as PDF on this dashboard for a formatted report.")
    # Print-ready summary table
    st.markdown("#### Print-Ready Risk Summary")
    summary_cols = ['risk_id', 'category', 'title', 'likelihood', 'impact', 'inherent_score', 'inherent_rating', 'control_effectiveness', 'residual_score', 'residual_rating', 'risk_owner', 'remediation_status', 'remediation_deadline']
    st.dataframe(df[summary_cols], use_container_width=True)

# END